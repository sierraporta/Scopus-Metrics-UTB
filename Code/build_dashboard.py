#!/usr/bin/env python3
"""
build_dashboard.py
==================
Genera el dashboard UTB Scopus (index.html) cruzando:
  - Export CSV de Scopus
  - Base de docentes de planta UTB (xlsx)
  - Scimago JR 2025 (csv) para cuartiles y áreas de revista

Ejecutar desde la carpeta ScriptScopusDef:
    python3 build_dashboard.py
"""
import pandas as pd
import numpy as np
import json, re, unicodedata
from pathlib import Path
from itertools import combinations
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ─── CONFIG ───────────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).parent
SCOPUS_CSV    = BASE_DIR / "scopus_export_Jun 14-2026_1d2db208-3a63-4746-be63-c20b3217c430.csv"
FACULTY_XLSX  = BASE_DIR / "Base de Datos Scopus 2024.xlsx"
SCIMAGO_CSV   = BASE_DIR / "scimagojr 2025.csv"
OUT_DIR       = BASE_DIR / "utb_scopus_dashboard_single_pretty"
START_YEAR    = 2022
TOP_AUTHORS   = 20
TOP_SCHOOLS   = 18
TOP_PAIRS     = 25
TOP_AREAS     = 15
DOC_TYPES     = ["Article", "Conference", "Review", "Other"]
QUARTILES     = ["Q1", "Q2", "Q3", "Q4", "No Q"]

# ─── HELPERS ──────────────────────────────────────────────────────────────────
_LOWER_ES = {"de","del","la","el","las","los","y","e","a","con","en","o","por","para","al"}

def title_case(s):
    if s is None or (isinstance(s, float) and np.isnan(s)): return s
    words = str(s).strip().split()
    return " ".join(
        w.lower() if i > 0 and w.lower() in _LOWER_ES else w.capitalize()
        for i, w in enumerate(words)
    )

def norm_issns(s):
    """Return list of normalised ISSNs (digits + X only, uppercase)."""
    if s is None or (isinstance(s, float) and np.isnan(s)): return []
    return [re.sub(r"[^0-9X]", "", x.upper()) for x in str(s).split(",") if x.strip()]

def doc_type_bucket(t):
    if t is None or (isinstance(t, float) and np.isnan(t)): return "Other"
    t = str(t).lower()
    if "review" in t:                          return "Review"
    if "conference" in t or "proceeding" in t: return "Conference"
    if "article" in t or "data paper" in t or "short survey" in t: return "Article"
    return "Other"

def _extract_authorid(url):
    if url is None or (isinstance(url, float) and np.isnan(url)): return None
    m = re.search(r"authorId=(\d+)", str(url))
    return m.group(1) if m else None

def _canon_id(v):
    if v is None or (isinstance(v, float) and np.isnan(v)): return None
    s = re.sub(r"[^0-9]", "", str(v))
    return s if s else None

def _split_sc(s):
    if s is None or (isinstance(s, float) and np.isnan(s)): return []
    return [x.strip() for x in str(s).split(";") if x.strip()]

def _parse_author_entry(entry):
    if not entry or (isinstance(entry, float) and np.isnan(entry)): return None, None
    parts = [p.strip() for p in str(entry).split(",", 2)]
    if len(parts) >= 2:
        return parts[0] + ", " + parts[1], parts[2] if len(parts) > 2 else ""
    return entry, ""

def safe_int(x):
    try: return int(x)
    except: return 0

# ─── LOAD FACULTY ─────────────────────────────────────────────────────────────
print("Loading faculty...")
faculty_raw = pd.read_excel(FACULTY_XLSX, sheet_name="Hoja1")
faculty_raw["_id_url"]  = faculty_raw.get("SCOPUS",   pd.Series([None]*len(faculty_raw))).apply(_extract_authorid)
faculty_raw["_id_cell"] = faculty_raw.get("ID SCOPUS",pd.Series([None]*len(faculty_raw))).apply(_canon_id)
faculty_raw["author_id"] = faculty_raw["_id_cell"].where(faculty_raw["_id_cell"].notna(), faculty_raw["_id_url"])
faculty_valid = faculty_raw[faculty_raw["author_id"].notna()].copy()
faculty_valid["author_id"] = faculty_valid["author_id"].astype(str).str.strip()
faculty_valid = faculty_valid.drop_duplicates("author_id").reset_index(drop=True)
faculty_valid["DOCENTE"] = faculty_valid["DOCENTE"].apply(title_case)
faculty_valid["ESCUELA"] = faculty_valid["ESCUELA"].apply(title_case)
faculty_ids  = set(faculty_valid["author_id"])
name_map     = faculty_valid.set_index("author_id")["DOCENTE"].to_dict()
school_map   = faculty_valid.set_index("author_id")["ESCUELA"].to_dict()
print(f"  Faculty loaded: {len(faculty_valid)} docentes")

# ─── LOAD SCOPUS ──────────────────────────────────────────────────────────────
print("Loading Scopus...")
df = pd.read_csv(SCOPUS_CSV, encoding="utf-8-sig")
df["Year"] = pd.to_numeric(df["Year"], errors="coerce")
df = df[df["Year"].notna() & (df["Year"] >= START_YEAR)].copy()
df["Year"] = df["Year"].astype(int)
df["doc_type3"] = df["Document Type"].apply(doc_type_bucket)
actu = SCOPUS_CSV.stem.split("_")[2]   # e.g. "Apr 15-2026"
print(f"  Scopus loaded: {len(df)} papers, years {df['Year'].min()}–{df['Year'].max()}")

# ─── LOAD SCIMAGO ─────────────────────────────────────────────────────────────
print("Loading Scimago...")
scimago = pd.read_csv(SCIMAGO_CSV, sep=";")
# Build ISSN → {quartile, sjr, title, categories, areas}
scimago_lookup = {}
for _, row in scimago.iterrows():
    q   = str(row.get("SJR Best Quartile","")).strip()
    sjr = str(row.get("SJR","0")).replace(",",".").strip()
    try: sjr_f = float(sjr)
    except: sjr_f = 0.0
    areas_raw = str(row.get("Areas","")).strip()
    areas_list = [a.strip() for a in areas_raw.split(";") if a.strip()] if areas_raw and areas_raw != "nan" else []
    info = {"quartile": q if q in ("Q1","Q2","Q3","Q4") else "No Q",
            "sjr": sjr_f,
            "scimago_title": str(row.get("Title","")),
            "categories": str(row.get("Categories","")),
            "areas": areas_list}
    for issn in norm_issns(row.get("Issn","")):
        scimago_lookup[issn] = info
print(f"  Scimago lookup: {len(scimago_lookup)} ISSNs")

def get_quartile(issn_raw):
    for issn in norm_issns(issn_raw):
        if issn in scimago_lookup:
            return scimago_lookup[issn]
    return {"quartile":"No Q","sjr":0.0,"scimago_title":"","categories":"","areas":[]}

# ─── BUILD UTB PAPER-LEVEL DATASET ────────────────────────────────────────────
print("Building author-paper dataset...")
records = []
for _, r in df.iterrows():
    eid   = r.get("EID")
    year  = int(r.get("Year"))
    dt3   = r.get("doc_type3")
    issn  = r.get("ISSN")
    src   = str(r.get("Source title",""))
    sci   = get_quartile(issn)

    ids   = _split_sc(r.get("Author(s) ID"))
    awas  = _split_sc(r.get("Authors with affiliations"))
    short = _split_sc(r.get("Authors"))
    n = max(len(ids), len(awas), len(short))
    ids  += [None]*(n-len(ids))
    awas += [None]*(n-len(awas))
    short+= [None]*(n-len(short))

    for i in range(n):
        name,_ = _parse_author_entry(awas[i]) if awas[i] else (None, None)
        if name is None and short[i]:
            name = str(short[i]).strip()
        aid = str(ids[i]).strip() if ids[i] else None
        if aid in ("","nan","None"): aid = None
        records.append({"EID":eid,"Year":year,"doc_type3":dt3,"ISSN":issn,
                        "Source":src,"quartile":sci["quartile"],"sjr":sci["sjr"],
                        "author_id":aid,"author_name":name})

authors_long = pd.DataFrame(records)
authors_long = authors_long.dropna(subset=["author_id","author_name"], how="all")
authors_long["author_id"] = authors_long["author_id"].astype(str).str.strip()

# Filter to UTB planta only
planta = authors_long[authors_long["author_id"].isin(faculty_ids)].copy()
planta["DOCENTE"] = planta["author_id"].map(name_map).fillna(planta["author_name"])
planta["ESCUELA"] = planta["author_id"].map(school_map).fillna("Sin Escuela")

# Unique-paper credit per (EID, ESCUELA) and per (EID, DOCENTE)
school_papers = planta[["EID","Year","ESCUELA","doc_type3","quartile","sjr"]].drop_duplicates()
author_papers = planta[["EID","Year","author_id","DOCENTE","ESCUELA","doc_type3","quartile","sjr"]].drop_duplicates()

# UTB unique papers for area expansion
utb_eids = set(school_papers["EID"].unique())
unique_utb = df[df["EID"].isin(utb_eids)][["EID","Year","doc_type3","ISSN","quartile"]
    if "quartile" in df.columns else ["EID","Year","doc_type3","ISSN"]].drop_duplicates("EID").copy()
# Re-attach quartile from planta (not df which has all papers)
utb_q = school_papers[["EID","Year","doc_type3","quartile"]].drop_duplicates("EID")

# Build EID → areas mapping from original df (via Scimago ISSN lookup)
eid_to_areas = {}
for _, r in df[df["EID"].isin(utb_eids)].iterrows():
    sci = get_quartile(r.get("ISSN"))
    areas = sci.get("areas", [])
    eid_to_areas[r["EID"]] = areas if areas else ["Sin clasificar"]

# Build area_papers: expand each UTB paper by its Scimago areas
area_rows = []
for _, r in utb_q.iterrows():
    for area in eid_to_areas.get(r["EID"], ["Sin clasificar"]):
        area_rows.append({"EID": r["EID"], "Year": r["Year"],
                          "doc_type3": r["doc_type3"], "quartile": r["quartile"],
                          "area": area})
area_papers = (pd.DataFrame(area_rows) if area_rows
               else pd.DataFrame(columns=["EID","Year","doc_type3","quartile","area"]))

years_list = sorted(df["Year"].unique().tolist())
year_sels  = ["ALL"] + [str(y) for y in years_list]
print(f"  UTB papers: {author_papers['EID'].nunique()} unique, "
      f"{author_papers['DOCENTE'].nunique()} authors, {author_papers['ESCUELA'].nunique()} schools")

# ─── AGGREGATION HELPERS ──────────────────────────────────────────────────────
def filter_year(df_in, ys):
    if ys == "ALL": return df_in
    return df_in[df_in["Year"] == int(ys)]

def doc_type_counts(base):
    return {dt: int(base[base["doc_type3"]==dt]["EID"].nunique()) for dt in DOC_TYPES}

def quartile_counts(base, articles_only=True):
    b = base[base["doc_type3"]=="Article"] if articles_only else base
    return {q: int(b[b["quartile"]==q]["EID"].nunique()) for q in QUARTILES}

def kpis(base_ap, base_sp):
    n_docs    = int(base_ap["EID"].nunique())
    arts      = base_ap[base_ap["doc_type3"]=="Article"]
    n_arts    = int(arts["EID"].nunique())
    q1        = int(arts[arts["quartile"]=="Q1"]["EID"].nunique())
    q1q2      = int(arts[arts["quartile"].isin(["Q1","Q2"])]["EID"].nunique())
    has_q     = int(arts[arts["quartile"]!="No Q"]["EID"].nunique())
    avg_sjr   = round(float(arts[arts["sjr"]>0]["sjr"].mean()),3) if arts[arts["sjr"]>0].shape[0]>0 else 0
    pct_q1    = round(q1/has_q*100,1)   if has_q > 0 else 0
    pct_q1q2  = round(q1q2/has_q*100,1) if has_q > 0 else 0
    n_authors = int(base_ap["author_id"].nunique())
    n_schools = int(base_sp["ESCUELA"].nunique())
    return dict(n_docs=n_docs, n_articles=n_arts, n_authors=n_authors,
                n_schools=n_schools, q1=q1, q1q2=q1q2, has_q=has_q,
                pct_q1=pct_q1, pct_q1q2=pct_q1q2, avg_sjr=avg_sjr)

def schools_data(base_ap, base_sp):
    tot = (base_sp[["EID","ESCUELA"]].drop_duplicates()
           .groupby("ESCUELA").size().reset_index(name="total")
           .sort_values("total", ascending=False))
    top = tot.head(TOP_SCHOOLS)["ESCUELA"].tolist()
    rows = []
    for esc in top:
        sub = base_sp[base_sp["ESCUELA"]==esc]
        art = sub[sub["doc_type3"]=="Article"]
        row = {"name": esc, "total": int(sub["EID"].nunique())}
        for dt in DOC_TYPES:
            row[dt] = int(sub[sub["doc_type3"]==dt]["EID"].nunique())
        for q in QUARTILES:
            row[q] = int(art[art["quartile"]==q]["EID"].nunique())
        has_q = sum(row[q] for q in ["Q1","Q2","Q3","Q4"])
        row["pct_q1"] = round(row["Q1"]/has_q*100, 1) if has_q > 0 else 0
        rows.append(row)
    return rows

def authors_data(base_ap):
    tot = (base_ap.drop_duplicates(["EID","author_id"])
           .groupby(["author_id","DOCENTE","ESCUELA"])["EID"].nunique()
           .reset_index(name="total")
           .sort_values("total", ascending=False)
           .head(TOP_AUTHORS))
    rows = []
    for _, r in tot.iterrows():
        sub = base_ap[base_ap["author_id"]==r["author_id"]].drop_duplicates("EID")
        art = sub[sub["doc_type3"]=="Article"]
        row = {"name": r["DOCENTE"], "school": r["ESCUELA"], "total": int(r["total"])}
        for dt in DOC_TYPES:
            row[dt] = int(sub[sub["doc_type3"]==dt].shape[0])
        for q in QUARTILES:
            row[q] = int(art[art["quartile"]==q].shape[0])
        has_q = sum(row[q] for q in ["Q1","Q2","Q3","Q4"])
        row["pct_q1"] = round(row["Q1"]/has_q*100,1) if has_q>0 else 0
        rows.append(row)
    return rows

def pairs_data(base_ap):
    pfac = base_ap[["EID","Year","author_id"]].drop_duplicates()
    pair_rows = []
    for eid, g in pfac.groupby("EID"):
        ids = sorted(set(g["author_id"].astype(str).tolist()))
        if len(ids) < 2: continue
        for a, b in combinations(ids, 2):
            pair_rows.append((a, b, eid))
    if not pair_rows:
        return []
    pairdf = pd.DataFrame(pair_rows, columns=["a","b","EID"])
    counts = (pairdf.groupby(["a","b"])
              .agg(n=("EID","nunique"))
              .reset_index()
              .sort_values("n", ascending=False)
              .head(TOP_PAIRS))
    rows = []
    for _, r in counts.iterrows():
        na = name_map.get(r["a"], r["a"])
        nb = name_map.get(r["b"], r["b"])
        rows.append({"pair": f"{na} — {nb}", "a": na, "b": nb, "n": int(r["n"])})
    return rows

def areas_data(area_sub):
    """Top areas by article count, with quartile breakdown. Each paper can appear in multiple areas."""
    art_sub = area_sub[area_sub["doc_type3"] == "Article"]
    if art_sub.empty:
        return []
    tot = (art_sub.groupby("area")["EID"].nunique()
           .reset_index(name="total")
           .sort_values("total", ascending=False)
           .head(TOP_AREAS))
    rows = []
    for _, r in tot.iterrows():
        sub = art_sub[art_sub["area"] == r["area"]]
        row = {"name": r["area"], "total": int(r["total"])}
        for q in QUARTILES:
            row[q] = int(sub[sub["quartile"]==q]["EID"].nunique())
        has_q = sum(row[q] for q in ["Q1","Q2","Q3","Q4"])
        row["pct_q1"] = round(row["Q1"] / has_q * 100, 1) if has_q > 0 else 0
        rows.append(row)
    return rows

def authors_pivot_data():
    """
    Pivot: one row per author, columns = (year × quartile) for ARTICLES only.
    Returns list of dicts sorted by school → total_arts desc.
    """
    arts   = author_papers[author_papers["doc_type3"] == "Article"].copy()
    all_ap = author_papers.copy()

    # Unique author info (pick first school/name per author_id)
    info_df = (author_papers
               .drop_duplicates("author_id")
               [["author_id","DOCENTE","ESCUELA"]]
               .copy())

    rows = []
    for _, info in info_df.iterrows():
        aid   = info["author_id"]
        a_art = arts[arts["author_id"] == aid]
        a_all = all_ap[all_ap["author_id"] == aid]
        row   = {"name": info["DOCENTE"],
                 "school": info["ESCUELA"],
                 "scopus_id": aid}
        grand_arts = 0
        for y in years_list:
            yg = a_art[a_art["Year"] == y]
            ya = a_all[a_all["Year"] == y]
            yd = {}
            for q in ["Q1","Q2","Q3","Q4"]:
                yd[q] = int(yg[yg["quartile"]==q]["EID"].nunique())
            yd["SC"]         = int(yg[yg["quartile"]=="No Q"]["EID"].nunique())
            yd["total_arts"] = int(yg["EID"].nunique())
            yd["total_docs"] = int(ya["EID"].nunique())
            row[str(y)] = yd
            grand_arts  += yd["total_arts"]
        row["_grand_arts"] = grand_arts
        rows.append(row)

    rows.sort(key=lambda r: (r["school"], -r["_grand_arts"]))
    # Remove helper key before serializing
    for r in rows:
        del r["_grand_arts"]
    return rows

# ─── BUILD TIMELINE (always all years, fixed) ─────────────────────────────────
print("Computing timeline aggregations...")
timeline = []
for y in years_list:
    sub = school_papers[school_papers["Year"]==y]
    row = {"year": str(y), "total": int(sub["EID"].nunique())}
    for dt in DOC_TYPES:
        row[dt] = int(sub[sub["doc_type3"]==dt]["EID"].nunique())
    timeline.append(row)

# Quartile trend: articles only, all years
q_trend = []
for y in years_list:
    sub = school_papers[(school_papers["Year"]==y) & (school_papers["doc_type3"]=="Article")]
    row = {"year": str(y), "total": int(sub["EID"].nunique())}
    for q in QUARTILES:
        row[q] = int(sub[sub["quartile"]==q]["EID"].nunique())
    q_trend.append(row)

# ─── BUILD BY-YEAR DATA ───────────────────────────────────────────────────────
print("Computing per-year slices...")
by_year = {}
for ys in year_sels:
    ap_sub   = filter_year(author_papers, ys)
    sp_sub   = filter_year(school_papers, ys)
    area_sub = filter_year(area_papers, ys)
    by_year[ys] = {
        "kpis":      kpis(ap_sub, sp_sub),
        "doc_types": doc_type_counts(sp_sub),
        "quartiles": quartile_counts(sp_sub),
        "schools":   schools_data(ap_sub, sp_sub),
        "authors":   authors_data(ap_sub),
        "pairs":     pairs_data(ap_sub),
        "areas":     areas_data(area_sub),
    }

# ─── FINAL PAYLOAD ────────────────────────────────────────────────────────────
payload = {
    "meta": {
        "updated":    actu,
        "start_year": START_YEAR,
        "years":      year_sels,
    },
    "timeline":       timeline,
    "quartile_trend": q_trend,
    "by_year":        by_year,
}

print("Building authors pivot...")
authors_pivot = authors_pivot_data()
print(f"  Pivot: {len(authors_pivot)} authors × {len(years_list)} years")

payload["authors_pivot"] = {"years": [str(y) for y in years_list],
                             "rows":  authors_pivot}
print("Data payload ready.")

# ─── HTML TEMPLATE ────────────────────────────────────────────────────────────
DATA_JSON = json.dumps(payload, ensure_ascii=False, separators=(",",":"))

HTML = r"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>UTB Scopus Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js"></script>
<style>
/* ── RESET & BASE ─────────────────────────────────────────────── */
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth;font-size:15px}
body{font-family:'Segoe UI',system-ui,-apple-system,Arial,sans-serif;
     background:#F0F4F9;color:#1E293B;-webkit-font-smoothing:antialiased}
/* ── HERO ─────────────────────────────────────────────────────── */
.hero{
  background:linear-gradient(135deg,#050E1F 0%,#0D2158 45%,#1A46CC 100%);
  padding:52px 72px 76px;position:relative;overflow:hidden;color:#fff}
.hero::before{content:'';position:absolute;inset:0;
  background-image:url("data:image/svg+xml,%3Csvg width='40' height='40' viewBox='0 0 40 40' xmlns='http://www.w3.org/2000/svg'%3E%3Cg fill='%23fff' fill-opacity='0.025'%3E%3Ccircle cx='20' cy='20' r='1.5'/%3E%3C/g%3E%3C/svg%3E")}
.hero-glow{position:absolute;top:-120px;right:-60px;width:520px;height:520px;
  border-radius:50%;background:radial-gradient(circle,rgba(100,130,255,.28) 0%,transparent 65%);pointer-events:none}
.hero-glow2{position:absolute;bottom:-80px;left:18%;width:320px;height:320px;
  border-radius:50%;background:radial-gradient(circle,rgba(20,180,220,.18) 0%,transparent 65%);pointer-events:none}
.hero-inner{position:relative;z-index:2;max-width:1180px;margin:0 auto;
  display:flex;align-items:flex-start;justify-content:space-between;gap:32px;flex-wrap:wrap}
.hero-text{flex:1;min-width:280px}
.hero-badge{display:inline-flex;align-items:center;gap:6px;
  background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.2);
  border-radius:999px;padding:4px 14px;font-size:11.5px;font-weight:600;
  letter-spacing:.3px;margin-bottom:18px;backdrop-filter:blur(6px)}
.hero-title{font-size:36px;font-weight:800;letter-spacing:-1px;line-height:1.1;margin-bottom:14px}
.hero-title span{color:#93C5FD}
.hero-desc{font-size:13.5px;line-height:1.8;color:rgba(255,255,255,.75);max-width:680px}
.hero-desc strong{color:rgba(255,255,255,.95)}
/* Year filter in hero */
.year-filter-wrap{display:flex;flex-direction:column;align-items:flex-end;gap:6px;flex-shrink:0;padding-top:4px}
.year-filter-label{font-size:11px;font-weight:700;text-transform:uppercase;
  letter-spacing:.8px;color:rgba(255,255,255,.6)}
#yearFilter{
  appearance:none;background:#1E3A8A url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 24 24' fill='none' stroke='%2393C5FD' stroke-width='2.5'%3E%3Cpath d='M6 9l6 6 6-6'/%3E%3C/svg%3E") no-repeat right 10px center;
  border:1.5px solid rgba(147,197,253,.4);border-radius:8px;
  color:#fff;font-size:14px;font-weight:600;padding:8px 36px 8px 14px;
  cursor:pointer;min-width:110px;transition:border-color .15s}
#yearFilter:hover{border-color:rgba(147,197,253,.75)}
/* ── KPI STRIP ────────────────────────────────────────────────── */
.kpi-strip{margin-bottom:24px;max-width:1180px;margin:-32px auto 0;padding:0 72px;position:relative;z-index:10}
.kpi-grid{display:grid;grid-template-columns:repeat(6,1fr);gap:13px}
.kpi-card{background:#fff;border-radius:13px;padding:17px 18px 14px;
  box-shadow:0 6px 28px rgba(15,23,42,.10);border-top:10px solid var(--ac);
  transition:transform .16s,box-shadow .16s;animation:riseIn .5s ease both}
.kpi-card:hover{transform:translateY(-3px);box-shadow:0 14px 40px rgba(15,23,42,.15)}
.kpi-card:nth-child(1){--ac:#2563EB;animation-delay:.04s}
.kpi-card:nth-child(2){--ac:#059669;animation-delay:.08s}
.kpi-card:nth-child(3){--ac:#10B981;animation-delay:.12s}
.kpi-card:nth-child(4){--ac:#6366F1;animation-delay:.16s}
.kpi-card:nth-child(5){--ac:#EA580C;animation-delay:.20s}
.kpi-card:nth-child(6){--ac:#0891B2;animation-delay:.24s}
.kpi-lbl{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;
  color:#94A3B8;margin-bottom:5px}
.kpi-val{font-size:28px;font-weight:800;line-height:1;color:#0F172A}
.kpi-val.sm{font-size:21px}
.kpi-sub{font-size:10.5px;color:#94A3B8;margin-top:4px}
/* ── NAV BAR ──────────────────────────────────────────────────── */
.nav-bar{margin-top:24px;position:sticky;top:0;z-index:50;background:rgba(255,255,255,.92);
  backdrop-filter:blur(10px);border-bottom:1px solid #E2E8F0;
  box-shadow:0 2px 12px rgba(15,23,42,.06)}
.nav-inner{max-width:1180px;margin:0 auto;padding:0 72px;
  display:flex;gap:0;overflow-x:auto}
.nav-link{padding:14px 20px;font-size:12.5px;font-weight:600;color:#64748B;
  text-decoration:none;border-bottom:2.5px solid transparent;white-space:nowrap;
  transition:color .15s,border-color .15s;letter-spacing:.1px}
.nav-link:hover{color:#1D4ED8}
.nav-link.active{color:#1D4ED8;border-color:#1D4ED8}
/* ── SECTIONS ─────────────────────────────────────────────────── */
.section{padding:52px 72px;max-width:1180px;margin:0 auto}
.section-hd{margin-bottom:28px}
.section-eye{font-size:10.5px;font-weight:700;text-transform:uppercase;
  letter-spacing:1px;color:#94A3B8;margin-bottom:6px}
.section-title{font-size:22px;font-weight:800;color:#0F172A;letter-spacing:-.4px}
.section-sub{font-size:13.5px;color:#64748B;margin-top:6px;line-height:1.6}
/* ── CHART CARDS ──────────────────────────────────────────────── */
.card{background:#fff;border-radius:16px;padding:28px 28px 20px;
  box-shadow:0 3px 20px rgba(15,23,42,.07);border:1px solid rgba(226,232,240,.8);
  margin-bottom:24px;animation:riseIn .5s ease both;
  transition:box-shadow .18s}
.card:hover{box-shadow:0 10px 40px rgba(15,23,42,.12)}
.card-hd{display:flex;align-items:center;justify-content:space-between;margin-bottom:20px}
.card-title{font-size:14px;font-weight:700;color:#1E293B;letter-spacing:-.1px}
.card-dot{width:8px;height:8px;border-radius:50%;display:inline-block;margin-right:7px}
.card-note{font-size:11.5px;color:#94A3B8;background:#F8FAFC;
  padding:4px 10px;border-radius:6px}
.card-hd-right{display:flex;align-items:center;gap:8px}
.chart-wrap{position:relative}
/* ── RESIZE CONTROLS ──────────────────────────────────────────── */
.resize-btn{
  background:none;border:1.5px solid #E2E8F0;border-radius:6px;
  cursor:pointer;color:#94A3B8;font-size:13px;padding:2px 8px;line-height:1.6;
  transition:border-color .15s,color .15s,background .15s;flex-shrink:0;
  font-family:inherit}
.resize-btn:hover{border-color:#94A3B8;color:#475569;background:#F8FAFC}
.resize-row{
  display:none;align-items:center;gap:12px;
  padding:12px 4px 2px;border-top:1px solid #F1F5F9;margin-top:16px}
.height-slider{
  flex:1;accent-color:#1D4ED8;cursor:pointer;
  -webkit-appearance:none;height:4px;background:#E2E8F0;border-radius:2px;outline:none}
.height-slider::-webkit-slider-thumb{
  -webkit-appearance:none;width:16px;height:16px;border-radius:50%;
  background:#1D4ED8;cursor:pointer;box-shadow:0 0 0 3px rgba(37,99,235,.15)}
.height-lbl{
  font-size:11.5px;color:#475569;min-width:44px;text-align:right;
  font-weight:600;font-variant-numeric:tabular-nums}
/* ── 2-COL GRID for small charts ─────────────────────────────── */
.grid-2{display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-bottom:24px}
/* ── NOTE BOXES ───────────────────────────────────────────────── */
.note{border-radius:9px;padding:13px 17px;font-size:13px;line-height:1.7;margin-bottom:18px}
.note-y{background:#FEFCE8;border-left:10px solid #F59E0B;color:#78350F}
.note-b{background:#EFF6FF;border-left:10px solid #3B82F6;color:#1E3A8A}
.note-s{background:#F8FAFC;border-left:10px solid #64748B;color:#334155}
.note ul{margin-left:16px;margin-top:6px}.note li{margin:4px 0}
/* ── CREDITS ──────────────────────────────────────────────────── */
.credits{background:#0F172A;border-radius:14px;padding:22px 26px;
  color:rgba(255,255,255,.7);font-size:13px;line-height:1.7;margin-bottom:24px}
.credits strong{color:#fff}
/* ── FOOTER ───────────────────────────────────────────────────── */
footer{background:#0F172A;color:rgba(255,255,255,.45);padding:28px 72px;
  font-size:12.5px;line-height:1.7;margin-top:40px}
footer strong{color:rgba(255,255,255,.75)}
/* ── DIVIDER ──────────────────────────────────────────────────── */
hr.div{border:none;border-top:1px solid #E2E8F0;margin:8px 0 28px}
/* ── PIVOT TABLE ──────────────────────────────────────────────── */
.pivot-controls{display:flex;align-items:center;gap:14px;margin-bottom:18px;flex-wrap:wrap}
.pivot-controls label{font-size:12px;font-weight:700;color:#64748B;text-transform:uppercase;letter-spacing:.5px}
.pivot-filter{appearance:none;padding:8px 36px 8px 14px;font-size:13px;font-weight:600;
  border:1.5px solid #E2E8F0;border-radius:8px;background:#fff url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='10' viewBox='0 0 24 24' fill='none' stroke='%2394A3B8' stroke-width='2.5'%3E%3Cpath d='M6 9l6 6 6-6'/%3E%3C/svg%3E") no-repeat right 10px center;
  color:#1E293B;cursor:pointer;transition:border-color .15s;font-family:inherit}
.pivot-filter:hover{border-color:#94A3B8}
.pivot-wrap{overflow-x:auto;border-radius:12px;border:1px solid #E2E8F0;
  box-shadow:0 2px 12px rgba(15,23,42,.06)}
.pivot-tbl{border-collapse:collapse;width:max-content;min-width:100%;font-size:12.5px}
.pivot-tbl th,.pivot-tbl td{padding:8px 11px;white-space:nowrap;
  border-bottom:1px solid #F1F5F9;border-right:1px solid #F1F5F9}
.pivot-tbl th{font-weight:700;color:#475569;background:#F8FAFC;
  position:sticky;top:0;z-index:2;text-align:center;font-size:11.5px}
.pivot-tbl thead tr:first-child th{top:0}
.pivot-tbl thead tr:nth-child(2) th{top:35px;border-top:2px solid #E2E8F0}
/* Sticky first 3 cols */
.c-fix-1{position:sticky!important;left:0;z-index:4!important;
  min-width:180px;max-width:180px;text-align:left!important}
.c-fix-2{position:sticky!important;left:180px;z-index:4!important;
  min-width:155px;max-width:155px;text-align:left!important;
  border-right:2px solid #E2E8F0!important}
.c-fix-3{position:sticky!important;left:335px;z-index:4!important;
  min-width:110px;max-width:110px;text-align:left!important;
  border-right:2px solid #E2E8F0!important}
.pivot-tbl thead .c-fix-1,.pivot-tbl thead .c-fix-2,.pivot-tbl thead .c-fix-3{z-index:6!important;background:#F8FAFC}
.pivot-tbl tbody tr:hover td{background:#F8FAFC}
.pivot-tbl tbody td{color:#334155;text-align:center}
.pivot-tbl tbody .c-fix-1,.pivot-tbl tbody .c-fix-2,.pivot-tbl tbody .c-fix-3{
  background:#fff;text-align:left;color:#0F172A}
/* School group header */
.school-hdr td{background:#EFF6FF!important;color:#1D4ED8;font-weight:700;
  font-size:12px;letter-spacing:.3px;border-top:2px solid #BFDBFE!important}
/* Cell colors */
.q1v{color:#059669;font-weight:700}
.q2v{color:#2563EB;font-weight:600}
.q3v{color:#D97706;font-weight:600}
.q4v{color:#DC2626;font-weight:600}
.scv{color:#94A3B8;font-weight:500}
.zerv{color:#E2E8F0}
.tot-art{color:#6366F1;font-weight:700}
.tot-doc{color:#0891B2;font-weight:600}
/* Year group separator */
.yr-sep{border-left:2px solid #E2E8F0!important}
/* Pivot stat row */
.pivot-stat{font-size:12px;color:#64748B;margin-top:10px;text-align:right}
/* ── DARK MODE ─────────────────────────────────────────────────── */
[data-theme="dark"]{color-scheme:dark}
[data-theme="dark"] body{background:#0F172A;color:#E2E8F0}
[data-theme="dark"] .card{background:#1E293B;border-color:rgba(255,255,255,.06);
  box-shadow:0 3px 20px rgba(0,0,0,.4)}
[data-theme="dark"] .card:hover{box-shadow:0 10px 40px rgba(0,0,0,.55)}
[data-theme="dark"] .nav-bar{background:rgba(15,23,42,.96);border-color:rgba(255,255,255,.08)}
[data-theme="dark"] .nav-link{color:#94A3B8}
[data-theme="dark"] .nav-link:hover{color:#93C5FD}
[data-theme="dark"] .nav-link.active{color:#93C5FD;border-color:#93C5FD}
[data-theme="dark"] .kpi-card{background:#1E293B;box-shadow:0 6px 28px rgba(0,0,0,.35)}
[data-theme="dark"] .kpi-val{color:#F1F5F9}
[data-theme="dark"] .kpi-sub,[data-theme="dark"] .kpi-lbl{color:#64748B}
[data-theme="dark"] .section-title{color:#F1F5F9}
[data-theme="dark"] .section-sub{color:#94A3B8}
[data-theme="dark"] .section-eye{color:#64748B}
[data-theme="dark"] .card-title{color:#E2E8F0}
[data-theme="dark"] .card-note{background:#0F172A;color:#64748B}
[data-theme="dark"] .note-y{background:#2D1D02;border-color:#D97706;color:#FDE68A}
[data-theme="dark"] .note-b{background:#0C1E3E;border-color:#3B82F6;color:#BAE6FD}
[data-theme="dark"] .note-s{background:#1E293B;border-color:#475569;color:#94A3B8}
[data-theme="dark"] hr.div{border-color:rgba(255,255,255,.08)}
[data-theme="dark"] .resize-btn{border-color:#334155;color:#64748B}
[data-theme="dark"] .resize-btn:hover{border-color:#475569;color:#94A3B8;background:#1E293B}
[data-theme="dark"] .resize-row{border-color:rgba(255,255,255,.06)}
[data-theme="dark"] .height-slider{background:#334155}
[data-theme="dark"] .height-lbl{color:#94A3B8}
[data-theme="dark"] .pivot-wrap{border-color:rgba(255,255,255,.08)}
[data-theme="dark"] .pivot-tbl th{background:#1E293B;color:#94A3B8}
[data-theme="dark"] .pivot-tbl th,[data-theme="dark"] .pivot-tbl td{border-color:rgba(255,255,255,.06)}
[data-theme="dark"] .pivot-tbl thead .c-fix-1,[data-theme="dark"] .pivot-tbl thead .c-fix-2,[data-theme="dark"] .pivot-tbl thead .c-fix-3{background:#1E293B}
[data-theme="dark"] .pivot-tbl tbody .c-fix-1,[data-theme="dark"] .pivot-tbl tbody .c-fix-2,[data-theme="dark"] .pivot-tbl tbody .c-fix-3{background:#0F172A}
[data-theme="dark"] .pivot-tbl tbody td{color:#94A3B8}
[data-theme="dark"] .pivot-tbl tbody tr:hover td{background:#1E293B}
[data-theme="dark"] .school-hdr td{background:#0C1E3E!important;color:#93C5FD;border-color:#1E3A8A!important}
[data-theme="dark"] .pivot-filter{background-color:#1E293B;border-color:#334155;color:#E2E8F0}
[data-theme="dark"] .zerv{color:#334155}
/* ── THEME TOGGLE BUTTON ───────────────────────────────────────── */
.theme-btn{
  margin-left:auto;padding:8px 16px;font-size:12.5px;font-weight:600;
  background:none;border:1.5px solid #E2E8F0;border-radius:8px;
  cursor:pointer;color:#64748B;display:flex;align-items:center;gap:7px;
  transition:border-color .15s,color .15s,background .15s;flex-shrink:0;
  font-family:inherit;align-self:center}
.theme-btn:hover{border-color:#94A3B8;color:#1E293B;background:#F8FAFC}
[data-theme="dark"] .theme-btn{border-color:#334155;color:#94A3B8}
[data-theme="dark"] .theme-btn:hover{border-color:#475569;color:#E2E8F0;background:#1E293B}
/* ── ANIMATIONS ───────────────────────────────────────────────── */
@keyframes riseIn{from{opacity:0;transform:translateY(18px)}to{opacity:1;transform:translateY(0)}}
/* ── RESPONSIVE ───────────────────────────────────────────────── */
@media(max-width:1000px){
  .hero{padding:40px 28px 70px}.kpi-strip{padding:0 28px}
  .kpi-grid{grid-template-columns:repeat(3,1fr)}
  .nav-inner,.section{padding-left:28px;padding-right:28px}
  .grid-2{grid-template-columns:1fr}}
@media(max-width:580px){
  .hero-title{font-size:26px}.kpi-grid{grid-template-columns:repeat(2,1fr)}
  .grid-2{grid-template-columns:1fr}}
</style>
</head>
<body>

<!-- HERO -->
<header class="hero">
  <div class="hero-glow"></div><div class="hero-glow2"></div>
  <div class="hero-inner">
    <div class="hero-text">
      <div class="hero-badge">&#128202; Scopus · Scimago JR 2025</div>
      <h1 class="hero-title">UTB Scopus Dashboard <span>&#8805; __START_YEAR__</span></h1>
      <p class="hero-desc">
        Caracterización bibliométrica de la producción científica de docentes de planta de la
        <strong>Universidad Tecnológica de Bolívar</strong>. Enriquecido con cuartiles
        <strong>Scimago JR 2025</strong> para medir calidad de revista.
        Actualizado: <strong>__ACTU__</strong>.
      </p>
    </div>
    <div class="year-filter-wrap">
      <span class="year-filter-label">Filtro de año</span>
      <select id="yearFilter"></select>
    </div>
  </div>
</header>

<!-- KPI STRIP -->
<div class="kpi-strip">
  <div class="kpi-grid">
    <div class="kpi-card"><div class="kpi-lbl">Documentos únicos</div>
      <div class="kpi-val" id="k-docs">—</div><div class="kpi-sub">desde __START_YEAR__</div></div>
    <div class="kpi-card"><div class="kpi-lbl">Artículos</div>
      <div class="kpi-val" id="k-arts">—</div><div class="kpi-sub">en revistas indexadas</div></div>
    <div class="kpi-card"><div class="kpi-lbl">% Q1</div>
      <div class="kpi-val" id="k-q1">—</div><div class="kpi-sub">de artículos con cuartil</div></div>
    <div class="kpi-card"><div class="kpi-lbl">% Q1+Q2</div>
      <div class="kpi-val" id="k-q1q2">—</div><div class="kpi-sub">de artículos con cuartil</div></div>
    <div class="kpi-card"><div class="kpi-lbl">Docentes activos</div>
      <div class="kpi-val" id="k-auth">—</div><div class="kpi-sub">con publicaciones</div></div>
    <div class="kpi-card"><div class="kpi-lbl">Escuelas</div>
      <div class="kpi-val" id="k-sch">—</div><div class="kpi-sub">con producción</div></div>
  </div>
</div>

<!-- NAV -->
<nav class="nav-bar" id="navBar">
  <div class="nav-inner">
    <a class="nav-link active" href="#sec-prod">📈 Producción</a>
    <a class="nav-link" href="#sec-calidad">🏆 Calidad (Scimago)</a>
    <a class="nav-link" href="#sec-escuelas">🏫 Escuelas</a>
    <a class="nav-link" href="#sec-autores">👤 Autores</a>
    <a class="nav-link" href="#sec-colab">🤝 Colaboración</a>
    <a class="nav-link" href="#sec-areas">🔬 Áreas</a>
    <a class="nav-link" href="#sec-tabla">📋 Tabla Autores</a>
    <a class="nav-link" href="#sec-metodo">📄 Metodología</a>
    <button class="theme-btn" id="themeBtn" onclick="toggleTheme()">🌙 Oscuro</button>
  </div>
</nav>

<!-- ══════════════════════════════════════════════════════════════ -->
<!-- SECCIÓN 1: PRODUCCIÓN GENERAL                                 -->
<!-- ══════════════════════════════════════════════════════════════ -->
<section class="section" id="sec-prod">
  <div class="section-hd">
    <div class="section-eye">Sección 1</div>
    <h2 class="section-title">Producción General</h2>
    <p class="section-sub">Evolución anual de documentos únicos clasificados por tipo.</p>
  </div>

  <div class="card">
    <div class="card-hd">
      <span><span class="card-dot" style="background:#2563EB"></span>
        <span class="card-title">Documentos por año — apilado por tipo</span></span>
      <span class="card-hd-right">
        <span class="card-note">Serie completa</span>
        <button class="resize-btn" onclick="toggleResize(this)" title="Ajustar altura">⇕</button>
      </span>
    </div>
    <div class="chart-wrap" style="height:200px"><canvas id="c-timeline"></canvas></div>
    <div class="resize-row">
      <span style="font-size:11px;color:#94A3B8;white-space:nowrap">Altura:</span>
      <input type="range" class="height-slider" min="150" max="1400" step="10" value="300" oninput="resizeChart(this)">
      <span class="height-lbl">200px</span>
    </div>
  </div>

  <div class="grid-2">
    <div class="card">
      <div class="card-hd">
        <span><span class="card-dot" style="background:#6366F1"></span>
          <span class="card-title">Composición por tipo</span></span>
        <span class="card-note" id="lbl-donut-type">ALL</span>
      </div>
      <div class="chart-wrap" style="height:240px"><canvas id="c-type-donut"></canvas></div>
    </div>
    <div class="card">
      <div class="card-hd">
        <span><span class="card-dot" style="background:#10B981"></span>
          <span class="card-title">Artículos en Q1 y Q2</span></span>
        <span class="card-note" id="lbl-q1q2-mini">ALL</span>
      </div>
      <div class="chart-wrap" style="height:240px"><canvas id="c-q1q2-mini"></canvas></div>
    </div>
  </div>
</section>

<!-- ══════════════════════════════════════════════════════════════ -->
<!-- SECCIÓN 2: CALIDAD (SCIMAGO)                                  -->
<!-- ══════════════════════════════════════════════════════════════ -->
<section class="section" id="sec-calidad">
  <div class="section-hd">
    <div class="section-eye">Sección 2</div>
    <h2 class="section-title">Calidad de Publicaciones — Scimago JR</h2>
    <p class="section-sub">
      Cuartiles asignados a los artículos mediante cruce por ISSN con Scimago JR 2025.
      Solo aplica a documentos tipo <em>Article</em>.
    </p>
  </div>

  <div class="card">
    <div class="card-hd">
      <span><span class="card-dot" style="background:#10B981"></span>
        <span class="card-title">Evolución de cuartiles por año — artículos</span></span>
      <span class="card-hd-right">
        <span class="card-note">Serie completa</span>
        <button class="resize-btn" onclick="toggleResize(this)" title="Ajustar altura">⇕</button>
      </span>
    </div>
    <div class="chart-wrap" style="height:200px"><canvas id="c-q-trend"></canvas></div>
    <div class="resize-row">
      <span style="font-size:11px;color:#94A3B8;white-space:nowrap">Altura:</span>
      <input type="range" class="height-slider" min="150" max="1400" step="10" value="300" oninput="resizeChart(this)">
      <span class="height-lbl">200px</span>
    </div>
  </div>

  <div class="grid-2">
    <div class="card">
      <div class="card-hd">
        <span><span class="card-dot" style="background:#10B981"></span>
          <span class="card-title">Distribución de cuartiles</span></span>
        <span class="card-note" id="lbl-q-donut">ALL</span>
      </div>
      <div class="chart-wrap" style="height:240px"><canvas id="c-q-donut"></canvas></div>
    </div>
    <div class="card">
      <div class="card-hd">
        <span><span class="card-dot" style="background:#6366F1"></span>
          <span class="card-title">% Artículos Q1 por escuela</span></span>
        <span class="card-hd-right">
          <span class="card-note" id="lbl-pct-q1">ALL</span>
          <button class="resize-btn" onclick="toggleResize(this)" title="Ajustar altura">⇕</button>
        </span>
      </div>
      <div class="chart-wrap" style="height:240px"><canvas id="c-pct-q1"></canvas></div>
      <div class="resize-row">
        <span style="font-size:11px;color:#94A3B8;white-space:nowrap">Altura:</span>
        <input type="range" class="height-slider" min="150" max="1400" step="10" value="240" oninput="resizeChart(this)">
        <span class="height-lbl">240px</span>
      </div>
    </div>
  </div>

  <div class="card">
    <div class="card-hd">
      <span><span class="card-dot" style="background:#10B981"></span>
        <span class="card-title">Artículos por escuela — apilado por cuartil</span></span>
      <span class="card-hd-right">
        <span class="card-note" id="lbl-sch-q">ALL</span>
        <button class="resize-btn" onclick="toggleResize(this)" title="Ajustar altura">⇕</button>
      </span>
    </div>
    <div class="chart-wrap" style="height:__SCH_H__px"><canvas id="c-sch-q"></canvas></div>
    <div class="resize-row">
      <span style="font-size:11px;color:#94A3B8;white-space:nowrap">Altura:</span>
      <input type="range" class="height-slider" min="150" max="1400" step="10" value="__SCH_H__" oninput="resizeChart(this)">
      <span class="height-lbl">__SCH_H__px</span>
    </div>
  </div>
</section>

<!-- ══════════════════════════════════════════════════════════════ -->
<!-- SECCIÓN 3: ESCUELAS                                           -->
<!-- ══════════════════════════════════════════════════════════════ -->
<section class="section" id="sec-escuelas">
  <div class="section-hd">
    <div class="section-eye">Sección 3</div>
    <h2 class="section-title">Producción por Escuela</h2>
    <p class="section-sub">Documentos únicos atribuidos a cada Escuela, apilados por tipo.</p>
  </div>
  <div class="card">
    <div class="card-hd">
      <span><span class="card-dot" style="background:#059669"></span>
        <span class="card-title">Documentos por escuela — apilado por tipo</span></span>
      <span class="card-hd-right">
        <span class="card-note" id="lbl-sch-type">ALL</span>
        <button class="resize-btn" onclick="toggleResize(this)" title="Ajustar altura">⇕</button>
      </span>
    </div>
    <div class="chart-wrap" style="height:__SCH_H__px"><canvas id="c-sch-type"></canvas></div>
    <div class="resize-row">
      <span style="font-size:11px;color:#94A3B8;white-space:nowrap">Altura:</span>
      <input type="range" class="height-slider" min="150" max="1400" step="10" value="__SCH_H__" oninput="resizeChart(this)">
      <span class="height-lbl">__SCH_H__px</span>
    </div>
  </div>
</section>

<!-- ══════════════════════════════════════════════════════════════ -->
<!-- SECCIÓN 4: AUTORES                                            -->
<!-- ══════════════════════════════════════════════════════════════ -->
<section class="section" id="sec-autores">
  <div class="section-hd">
    <div class="section-eye">Sección 4</div>
    <h2 class="section-title">Top Autores</h2>
    <p class="section-sub">Los __TOP_AUTHORS__ docentes con mayor producción.</p>
  </div>

  <div class="card">
    <div class="card-hd">
      <span><span class="card-dot" style="background:#EA580C"></span>
        <span class="card-title">Top __TOP_AUTHORS__ autores — documentos por tipo</span></span>
      <span class="card-hd-right">
        <span class="card-note" id="lbl-auth-type">ALL</span>
        <button class="resize-btn" onclick="toggleResize(this)" title="Ajustar altura">⇕</button>
      </span>
    </div>
    <div class="chart-wrap" style="height:__AUTH_H__px"><canvas id="c-auth-type"></canvas></div>
    <div class="resize-row">
      <span style="font-size:11px;color:#94A3B8;white-space:nowrap">Altura:</span>
      <input type="range" class="height-slider" min="150" max="1400" step="10" value="__AUTH_H__" oninput="resizeChart(this)">
      <span class="height-lbl">__AUTH_H__px</span>
    </div>
  </div>

  <div class="card">
    <div class="card-hd">
      <span><span class="card-dot" style="background:#10B981"></span>
        <span class="card-title">Top __TOP_AUTHORS__ autores — artículos Q1</span></span>
      <span class="card-hd-right">
        <span class="card-note" id="lbl-auth-q1">ALL</span>
        <button class="resize-btn" onclick="toggleResize(this)" title="Ajustar altura">⇕</button>
      </span>
    </div>
    <div class="chart-wrap" style="height:__AUTH_H__px"><canvas id="c-auth-q1"></canvas></div>
    <div class="resize-row">
      <span style="font-size:11px;color:#94A3B8;white-space:nowrap">Altura:</span>
      <input type="range" class="height-slider" min="150" max="1400" step="10" value="__AUTH_H__" oninput="resizeChart(this)">
      <span class="height-lbl">__AUTH_H__px</span>
    </div>
  </div>
</section>

<!-- ══════════════════════════════════════════════════════════════ -->
<!-- SECCIÓN 5: COLABORACIÓN                                       -->
<!-- ══════════════════════════════════════════════════════════════ -->
<section class="section" id="sec-colab">
  <div class="section-hd">
    <div class="section-eye">Sección 5</div>
    <h2 class="section-title">Redes de Colaboración</h2>
    <p class="section-sub">Top __TOP_PAIRS__ pares de coautoría entre docentes de planta UTB.</p>
  </div>
  <div class="card">
    <div class="card-hd">
      <span><span class="card-dot" style="background:#0891B2"></span>
        <span class="card-title">Pares más frecuentes — documentos compartidos</span></span>
      <span class="card-hd-right">
        <span class="card-note" id="lbl-pairs">ALL</span>
        <button class="resize-btn" onclick="toggleResize(this)" title="Ajustar altura">⇕</button>
      </span>
    </div>
    <div class="chart-wrap" style="height:__PAIRS_H__px"><canvas id="c-pairs"></canvas></div>
    <div class="resize-row">
      <span style="font-size:11px;color:#94A3B8;white-space:nowrap">Altura:</span>
      <input type="range" class="height-slider" min="150" max="1400" step="10" value="__PAIRS_H__" oninput="resizeChart(this)">
      <span class="height-lbl">__PAIRS_H__px</span>
    </div>
  </div>
</section>

<!-- ══════════════════════════════════════════════════════════════ -->
<!-- SECCIÓN 6: ÁREAS SCIMAGO                                      -->
<!-- ══════════════════════════════════════════════════════════════ -->
<section class="section" id="sec-areas">
  <div class="section-hd">
    <div class="section-eye">Sección 6</div>
    <h2 class="section-title">Áreas Temáticas — Scimago JR</h2>
    <p class="section-sub">
      Distribución de artículos por área temática Scimago (top __TOP_AREAS__ áreas).
      Un artículo puede aparecer en múltiples áreas si la revista está clasificada en varias.
      Solo aplica a documentos tipo <em>Article</em>.
    </p>
  </div>

  <!-- Stacked bar: articles per area × quartile -->
  <div class="card">
    <div class="card-hd">
      <span><span class="card-dot" style="background:#10B981"></span>
        <span class="card-title">Artículos por área — apilado por cuartil</span></span>
      <span class="card-hd-right">
        <span class="card-note" id="lbl-area-q">ALL</span>
        <button class="resize-btn" onclick="toggleResize(this)" title="Ajustar altura">⇕</button>
      </span>
    </div>
    <div class="chart-wrap" style="height:__AREA_H__px"><canvas id="c-area-q"></canvas></div>
    <div class="resize-row">
      <span style="font-size:11px;color:#94A3B8;white-space:nowrap">Altura:</span>
      <input type="range" class="height-slider" min="150" max="1400" step="10" value="__AREA_H__" oninput="resizeChart(this)">
      <span class="height-lbl">__AREA_H__px</span>
    </div>
  </div>

  <!-- % Q1 per area (traffic-light) -->
  <div class="card">
    <div class="card-hd">
      <span><span class="card-dot" style="background:#6366F1"></span>
        <span class="card-title">% Artículos Q1 por área temática</span></span>
      <span class="card-hd-right">
        <span class="card-note" id="lbl-area-pct">ALL</span>
        <button class="resize-btn" onclick="toggleResize(this)" title="Ajustar altura">⇕</button>
      </span>
    </div>
    <div class="chart-wrap" style="height:__AREA_H__px"><canvas id="c-area-pct"></canvas></div>
    <div class="resize-row">
      <span style="font-size:11px;color:#94A3B8;white-space:nowrap">Altura:</span>
      <input type="range" class="height-slider" min="150" max="1400" step="10" value="__AREA_H__" oninput="resizeChart(this)">
      <span class="height-lbl">__AREA_H__px</span>
    </div>
  </div>
</section>

<!-- ══════════════════════════════════════════════════════════════ -->
<!-- SECCIÓN 7: TABLA PIVOT AUTORES                                -->
<!-- ══════════════════════════════════════════════════════════════ -->
<section class="section" id="sec-tabla">
  <div class="section-hd">
    <div class="section-eye">Sección 7</div>
    <h2 class="section-title">Tabla Pivot — Artículos por Autor y Cuartil</h2>
    <p class="section-sub">
      Un artículo por <code>EID</code> único. Columnas Q1–Q4 y SC (sin cuartil asignado)
      muestran el total de <em>artículos</em> por año. Las últimas columnas (Arts / Docs)
      incluyen todos los tipos de documento.
    </p>
  </div>

  <div class="card" style="padding:22px 22px 18px">
    <div class="pivot-controls">
      <label for="pivotSchool">Escuela:</label>
      <select id="pivotSchool" class="pivot-filter" onchange="renderPivot()">
        <option value="">Todas las escuelas</option>
      </select>
      <span class="pivot-stat" id="pivot-stat" style="margin-left:auto"></span>
    </div>
    <div class="pivot-wrap">
      <table class="pivot-tbl" id="pivotTable">
        <thead id="pivotHead"></thead>
        <tbody id="pivotBody"></tbody>
      </table>
    </div>
    <p style="font-size:11.5px;color:#94A3B8;margin-top:12px">
      SC = sin cuartil asignado en Scimago JR 2025 &nbsp;·&nbsp;
      Arts = total artículos &nbsp;·&nbsp; Docs = total documentos (todos los tipos)
    </p>
  </div>
</section>

<!-- ══════════════════════════════════════════════════════════════ -->
<!-- SECCIÓN 8: METODOLOGÍA                                        -->
<!-- ══════════════════════════════════════════════════════════════ -->
<section class="section" id="sec-metodo">
  <div class="section-hd">
    <div class="section-eye">Sección 8</div>
    <h2 class="section-title">Metodología y Notas</h2>
  </div>

  <div class="note note-b">
    <strong>Fuentes de datos:</strong>
    <ul>
      <li>Export CSV de Scopus (EID, año, tipo, autores, ISSN) filtrado desde __START_YEAR__.</li>
      <li>Base maestra de docentes de planta UTB con <em>Scopus Author ID</em>.</li>
      <li><strong>Scimago JR 2025</strong>: cruce por ISSN normalizado para asignar cuartil (Q1–Q4) y área temática a cada artículo.</li>
    </ul>
  </div>

  <div class="note note-y">
    <strong>Notas metodológicas:</strong>
    <ul>
      <li><strong>Unidad de conteo:</strong> documentos únicos por <code>EID</code>, no apariciones de autor.</li>
      <li><strong>Crédito por Escuela:</strong> una Escuela recibe crédito si al menos un docente de esa Escuela es autor (un documento puede contar en varias Escuelas).</li>
      <li><strong>Cuartiles y áreas:</strong> solo aplican a artículos (<em>Article</em>). Conference papers, libros, etc. aparecen como "No Q" / "Sin clasificar".</li>
      <li><strong>Áreas múltiples:</strong> si una revista pertenece a varias áreas Scimago, el artículo se contabiliza en cada área (doble conteo intencional).</li>
      <li><strong>% Q1 / Q1+Q2:</strong> calculado sobre artículos con cuartil asignado (excluye "No Q").</li>
    </ul>
  </div>

  <div class="note note-s">
    <strong>Alcance e interpretación:</strong>
    Este tablero es un ejercicio técnico de análisis bibliométrico basado en un export puntual de Scopus.
    Los resultados son <em>referenciales</em> y pueden diferir de cifras institucionales oficiales.
    No constituye un reporte oficial ni representa una posición institucional de la UTB.
  </div>

  <a class="card" href="tables.xlsx" style="display:flex;align-items:center;gap:18px;
    text-decoration:none;color:inherit;border-left:4px solid #059669">
    <span style="font-size:32px">📊</span>
    <div>
      <div style="font-weight:700;color:#0F172A;margin-bottom:3px">Tablas de datos (Excel)</div>
      <div style="font-size:13px;color:#64748B">
        Archivo <code>tables.xlsx</code> con tablas intermedias para auditoría y reportes.
      </div>
    </div>
  </a>

  <a class="card" href="autores_pivot_cuartiles.xlsx" style="display:flex;align-items:center;gap:18px;
    text-decoration:none;color:inherit;border-left:4px solid #6366F1">
    <span style="font-size:32px">🗂️</span>
    <div>
      <div style="font-weight:700;color:#0F172A;margin-bottom:3px">Tabla pivot autores — cuartiles por año (Excel)</div>
      <div style="font-size:13px;color:#64748B">
        Archivo <code>autores_pivot_cuartiles.xlsx</code> con Q1·Q2·Q3·Q4·SC por autor y año,
        agrupado por escuela. Paneles congelados listos para Excel.
      </div>
    </div>
  </a>

  <div class="credits">
    <strong>Créditos</strong> &mdash;
    Desarrollado por <strong>D. Sierra-Porta</strong> &copy; 2026 &middot;
    Universidad Tecnológica de Bolívar &middot;
    <em>Datos: Scopus + Scimago JR 2025 &middot; Actualizado: __ACTU__</em>
  </div>
</section>

<footer>
  <strong>© 2026 D. Sierra-Porta — UTB</strong> &nbsp;·&nbsp;
  Período: desde __START_YEAR__ &nbsp;·&nbsp;
  Fuente: Scopus + Scimago JR 2025 &nbsp;·&nbsp;
  Conteos sobre documentos únicos (EID)
</footer>

<script>
// ── DATA ────────────────────────────────────────────────────────
const D = __DATA_JSON__;

// ── PALETTE ─────────────────────────────────────────────────────
const C = {
  Article:'#2563EB', Conference:'#EA580C', Review:'#059669', Other:'#9CA3AF',
  Q1:'#10B981', Q2:'#3B82F6', Q3:'#F59E0B', Q4:'#EF4444', 'No Q':'#CBD5E1',
};
const TYPE_LABELS = ['Article','Conference','Review','Other'];
const Q_LABELS    = ['Q1','Q2','Q3','Q4','No Q'];
const FONT = "'Segoe UI','Inter',Arial,sans-serif";

// Chart.js defaults
Chart.defaults.font.family = FONT;
Chart.defaults.color = '#64748B';
Chart.defaults.plugins.legend.labels.boxWidth = 11;
Chart.defaults.plugins.legend.labels.padding  = 14;

// ── STATE ────────────────────────────────────────────────────────
let year = 'ALL';
const charts = {};

// ── UTILITIES ────────────────────────────────────────────────────
function mkChart(id, cfg){
  const ctx = document.getElementById(id);
  if(!ctx) return;
  if(charts[id]) charts[id].destroy();
  charts[id] = new Chart(ctx, cfg);
}
function hBar(labels, datasets, opts={}){
  return {type:'bar',
    data:{labels, datasets},
    options:{
      indexAxis:'y', responsive:true, maintainAspectRatio:false,
      animation:{duration:420, easing:'easeOutQuart'},
      plugins:{
        legend:{position:'top', display:opts.legend!==false},
        tooltip:{callbacks:{label: ctx => ` ${ctx.dataset.label||''}: ${ctx.parsed.x}`}}
      },
      scales:{
        x:{stacked:!!opts.stacked, grid:{color:'rgba(0,0,0,0.05)'},
           ticks:{font:{size:11}}},
        y:{stacked:!!opts.stacked, grid:{display:false},
           ticks:{font:{size:11}, maxRotation:0}}
      },
      ...opts.extra
    }
  };
}
function doughnut(labels, data, colors, opts={}){
  const total = data.reduce((a,b)=>a+b,0);
  return {type:'doughnut',
    data:{labels, datasets:[{data, backgroundColor:colors,
      borderWidth:2, borderColor:'#fff', hoverBorderColor:'#fff'}]},
    options:{
      responsive:true, maintainAspectRatio:false, cutout:'68%',
      animation:{duration:420},
      plugins:{
        legend:{position:'right', labels:{font:{size:11}, padding:10}},
        tooltip:{callbacks:{label: ctx => {
          const pct = total>0?(ctx.parsed/total*100).toFixed(1):'0';
          return ` ${ctx.label}: ${ctx.parsed} (${pct}%)`;
        }}}
      },
      ...opts
    }
  };
}
function label(id, text){ const el=document.getElementById(id); if(el) el.textContent=text; }

// ── DARK / LIGHT THEME ───────────────────────────────────────────
let darkMode = localStorage.getItem('utb-theme') === 'dark';
function applyTheme(dark){
  document.documentElement.setAttribute('data-theme', dark ? 'dark' : 'light');
  const btn = document.getElementById('themeBtn');
  if(btn) btn.innerHTML = dark ? '☀️ Claro' : '🌙 Oscuro';
  // Update Chart.js color defaults
  const col  = dark ? '#94A3B8' : '#64748B';
  const grid = dark ? 'rgba(255,255,255,0.05)' : 'rgba(0,0,0,0.05)';
  Chart.defaults.color = col;
  // Re-apply to existing charts
  Object.values(charts).forEach(ch=>{
    if(ch.config && ch.config.type === 'bar'){
      ['x','y'].forEach(ax=>{
        if(ch.options.scales[ax]){
          if(ch.options.scales[ax].grid) ch.options.scales[ax].grid.color = grid;
          if(ch.options.scales[ax].ticks) ch.options.scales[ax].ticks.color = col;
        }
      });
    }
    if(ch.options.plugins && ch.options.plugins.legend && ch.options.plugins.legend.labels)
      ch.options.plugins.legend.labels.color = col;
    ch.update('none');
  });
}
function toggleTheme(){
  darkMode = !darkMode;
  localStorage.setItem('utb-theme', darkMode ? 'dark' : 'light');
  applyTheme(darkMode);
}

// ── HEIGHT RESIZE CONTROLS ───────────────────────────────────────
function toggleResize(btn){
  const card = btn.closest('.card');
  const row  = card.querySelector('.resize-row');
  if(!row) return;
  const visible = row.style.display === 'flex';
  row.style.display = visible ? 'none' : 'flex';
  btn.style.background = visible ? '' : '#EFF6FF';
  btn.style.color      = visible ? '' : '#1D4ED8';
  btn.style.borderColor= visible ? '' : '#93C5FD';
}
function resizeChart(slider){
  const card   = slider.closest('.card');
  const wrap   = card.querySelector('.chart-wrap');
  const lbl    = card.querySelector('.height-lbl');
  const h      = parseInt(slider.value);
  if(wrap) wrap.style.height = h + 'px';
  if(lbl)  lbl.textContent   = h + 'px';
  const canvas = wrap ? wrap.querySelector('canvas') : null;
  if(canvas && charts[canvas.id]) charts[canvas.id].resize();
}

// ── KPI UPDATE ───────────────────────────────────────────────────
function updateKPIs(){
  const k = D.by_year[year].kpis;
  document.getElementById('k-docs').textContent  = k.n_docs;
  document.getElementById('k-arts').textContent  = k.n_articles;
  document.getElementById('k-q1').textContent    = k.pct_q1 + '%';
  document.getElementById('k-q1q2').textContent  = k.pct_q1q2 + '%';
  document.getElementById('k-auth').textContent  = k.n_authors;
  document.getElementById('k-sch').textContent   = k.n_schools;
}

// ── CHART 1: TIMELINE (fixed, all years) ────────────────────────
function drawTimeline(){
  const tl = D.timeline;
  const labels   = tl.map(r=>r.year);
  const datasets = TYPE_LABELS.map(t=>({
    label:t, data:tl.map(r=>r[t]||0),
    backgroundColor:C[t], borderRadius:4, borderSkipped:false,
  }));
  mkChart('c-timeline', hBar(labels, datasets, {stacked:true}));
}

// ── CHART 2: TYPE DONUT ─────────────────────────────────────────
function drawTypeDonut(){
  const dt   = D.by_year[year].doc_types;
  const vals = TYPE_LABELS.map(t=>dt[t]||0);
  label('lbl-donut-type', year);
  mkChart('c-type-donut', doughnut(TYPE_LABELS, vals, TYPE_LABELS.map(t=>C[t])));
}

// ── CHART 3: Q1+Q2 mini donut ───────────────────────────────────
function drawQ1Q2Mini(){
  const k   = D.by_year[year].kpis;
  const q1  = k.q1  || 0;
  const q2  = (D.by_year[year].quartiles['Q2']||0);
  const q3  = (D.by_year[year].quartiles['Q3']||0);
  const q4  = (D.by_year[year].quartiles['Q4']||0);
  const noq = (D.by_year[year].quartiles['No Q']||0);
  label('lbl-q1q2-mini', year);
  mkChart('c-q1q2-mini', doughnut(
    ['Q1','Q2','Q3','Q4','No Q'],
    [q1,q2,q3,q4,noq],
    [C.Q1,C.Q2,C.Q3,C.Q4,C['No Q']]
  ));
}

// ── CHART 4: QUARTILE TREND (fixed, all years) ──────────────────
function drawQTrend(){
  const qt = D.quartile_trend;
  const labels   = qt.map(r=>r.year);
  const datasets = Q_LABELS.map(q=>({
    label:q, data:qt.map(r=>r[q]||0),
    backgroundColor:C[q], borderRadius:3, borderSkipped:false,
  }));
  mkChart('c-q-trend', hBar(labels, datasets, {stacked:true}));
}

// ── CHART 5: QUARTILE DONUT ─────────────────────────────────────
function drawQDonut(){
  const q    = D.by_year[year].quartiles;
  const vals = Q_LABELS.map(l=>q[l]||0);
  label('lbl-q-donut', year);
  mkChart('c-q-donut', doughnut(Q_LABELS, vals, Q_LABELS.map(l=>C[l])));
}

// ── CHART 6: % Q1 POR ESCUELA ───────────────────────────────────
function drawPctQ1(){
  const schools = D.by_year[year].schools.slice(0,10);
  const labels  = schools.map(s=>s.name);
  const vals    = schools.map(s=>s.pct_q1||0);
  label('lbl-pct-q1', year);
  mkChart('c-pct-q1',{type:'bar',
    data:{labels, datasets:[{
      label:'% Q1', data:vals,
      backgroundColor:schools.map(s=>
        s.pct_q1>=60?'#10B981':s.pct_q1>=40?'#3B82F6':s.pct_q1>=20?'#F59E0B':'#EF4444'),
      borderRadius:4, borderSkipped:false,
    }]},
    options:{
      indexAxis:'y', responsive:true, maintainAspectRatio:false,
      animation:{duration:420},
      plugins:{legend:{display:false},
        tooltip:{callbacks:{label:ctx=>` ${ctx.parsed.x.toFixed(1)}% artículos en Q1`}}},
      scales:{
        x:{max:100, grid:{color:'rgba(0,0,0,0.05)'},
           ticks:{callback:v=>v+'%', font:{size:11}}},
        y:{grid:{display:false}, ticks:{font:{size:11}}}
      }
    }
  });
}

// ── CHART 7: SCHOOL × QUARTILE ──────────────────────────────────
function drawSchoolQ(){
  const schools  = D.by_year[year].schools;
  const labels   = schools.map(s=>s.name).reverse();
  label('lbl-sch-q', year);
  const datasets = Q_LABELS.map(q=>({
    label:q, data:schools.map(s=>s[q]||0).reverse(),
    backgroundColor:C[q], borderRadius:3, borderSkipped:false,
  }));
  mkChart('c-sch-q', hBar(labels, datasets, {stacked:true}));
}

// ── CHART 8: SCHOOL × TYPE ──────────────────────────────────────
function drawSchoolType(){
  const schools  = D.by_year[year].schools;
  const labels   = schools.map(s=>s.name).reverse();
  label('lbl-sch-type', year);
  const datasets = TYPE_LABELS.map(t=>({
    label:t, data:schools.map(s=>s[t]||0).reverse(),
    backgroundColor:C[t], borderRadius:3, borderSkipped:false,
  }));
  mkChart('c-sch-type', hBar(labels, datasets, {stacked:true}));
}

// ── CHART 9: AUTHORS × TYPE ─────────────────────────────────────
function drawAuthType(){
  const authors  = D.by_year[year].authors;
  const labels   = authors.map(a=>a.name).reverse();
  label('lbl-auth-type', year);
  const datasets = TYPE_LABELS.map(t=>({
    label:t, data:authors.map(a=>a[t]||0).reverse(),
    backgroundColor:C[t], borderRadius:3, borderSkipped:false,
  }));
  mkChart('c-auth-type', hBar(labels, datasets, {stacked:true}));
}

// ── CHART 10: AUTHORS × Q1 ──────────────────────────────────────
function drawAuthQ1(){
  const authors = D.by_year[year].authors
    .filter(a=>(a.Q1||0)>0)
    .sort((a,b)=>(b.Q1||0)-(a.Q1||0))
    .slice(0,15);
  const labels = authors.map(a=>a.name).reverse();
  label('lbl-auth-q1', year);
  mkChart('c-auth-q1',{type:'bar',
    data:{labels, datasets:[{
      label:'Artículos Q1', data:authors.map(a=>a.Q1||0).reverse(),
      backgroundColor:'#10B981', borderRadius:4, borderSkipped:false,
    }]},
    options:{
      indexAxis:'y', responsive:true, maintainAspectRatio:false,
      animation:{duration:420},
      plugins:{legend:{display:false},
        tooltip:{callbacks:{label:ctx=>` ${ctx.parsed.x} artículos Q1`}}},
      scales:{
        x:{grid:{color:'rgba(0,0,0,0.05)'}, ticks:{font:{size:11}}},
        y:{grid:{display:false}, ticks:{font:{size:11}}}
      }
    }
  });
}

// ── CHART 11: PAIRS ─────────────────────────────────────────────
function drawPairs(){
  const pairs = D.by_year[year].pairs;
  if(!pairs.length){ label('lbl-pairs', year+' (sin datos)'); return; }
  const labels = pairs.map(p=>p.pair).reverse();
  const vals   = pairs.map(p=>p.n).reverse();
  label('lbl-pairs', year);
  const maxN   = Math.max(...vals, 1);
  const colors = vals.map(v=>{
    const t = v/maxN;
    return `rgb(${Math.round(191-t*110)},${Math.round(219-t*110)},${Math.round(254-t*60)})`;
  });
  mkChart('c-pairs',{type:'bar',
    data:{labels, datasets:[{
      label:'Docs compartidos', data:vals,
      backgroundColor:colors, borderRadius:4, borderSkipped:false,
    }]},
    options:{
      indexAxis:'y', responsive:true, maintainAspectRatio:false,
      animation:{duration:420},
      plugins:{legend:{display:false},
        tooltip:{callbacks:{label:ctx=>` ${ctx.parsed.x} documentos compartidos`}}},
      scales:{
        x:{grid:{color:'rgba(0,0,0,0.05)'}, ticks:{font:{size:11}}},
        y:{grid:{display:false}, ticks:{font:{size:11}}}
      }
    }
  });
}

// ── CHART 12: AREAS × QUARTILE ──────────────────────────────────
function drawAreaQ(){
  const areas = D.by_year[year].areas;
  if(!areas || !areas.length){ label('lbl-area-q', year+' (sin datos)'); return; }
  const labels   = areas.map(a=>a.name).reverse();
  label('lbl-area-q', year);
  const datasets = Q_LABELS.map(q=>({
    label:q, data:areas.map(a=>a[q]||0).reverse(),
    backgroundColor:C[q], borderRadius:3, borderSkipped:false,
  }));
  mkChart('c-area-q', hBar(labels, datasets, {stacked:true}));
}

// ── CHART 13: % Q1 POR ÁREA ─────────────────────────────────────
function drawAreaPct(){
  const areas = D.by_year[year].areas;
  if(!areas || !areas.length){ label('lbl-area-pct', year+' (sin datos)'); return; }
  const sorted = [...areas].sort((a,b)=>(b.pct_q1||0)-(a.pct_q1||0));
  const labels = sorted.map(a=>a.name).reverse();
  const vals   = sorted.map(a=>a.pct_q1||0).reverse();
  label('lbl-area-pct', year);
  mkChart('c-area-pct',{type:'bar',
    data:{labels, datasets:[{
      label:'% Q1', data:vals,
      backgroundColor: vals.map(v=>
        v>=60?'#10B981':v>=40?'#3B82F6':v>=20?'#F59E0B':'#EF4444'),
      borderRadius:4, borderSkipped:false,
    }]},
    options:{
      indexAxis:'y', responsive:true, maintainAspectRatio:false,
      animation:{duration:420},
      plugins:{legend:{display:false},
        tooltip:{callbacks:{label:ctx=>` ${ctx.parsed.x.toFixed(1)}% artículos en Q1`}}},
      scales:{
        x:{max:100, grid:{color:'rgba(0,0,0,0.05)'},
           ticks:{callback:v=>v+'%', font:{size:11}}},
        y:{grid:{display:false}, ticks:{font:{size:11}}}
      }
    }
  });
}

// ── PIVOT TABLE ─────────────────────────────────────────────────
const P = D.authors_pivot;
const Q_COLS = ['Q1','Q2','Q3','Q4','SC'];
const Q_CSS  = {Q1:'q1v',Q2:'q2v',Q3:'q3v',Q4:'q4v',SC:'scv'};

function buildPivotHeader(){
  const head = document.getElementById('pivotHead');
  if(!head) return;
  // Row 1: fixed cols + year groups + total
  let r1 = '<tr>';
  r1 += `<th class="c-fix-1" rowspan="2">Docente</th>`;
  r1 += `<th class="c-fix-2" rowspan="2">Escuela</th>`;
  r1 += `<th class="c-fix-3" rowspan="2">Scopus ID</th>`;
  P.years.forEach((y,i)=>{
    const cls = i===0?'yr-sep':'';
    r1 += `<th colspan="7" style="text-align:center;background:#F0F4F9;${i>0?'border-left:2px solid #E2E8F0':''}">${y}</th>`;
  });
  r1 += `<th colspan="2" style="background:#EFF6FF;border-left:2px solid #C7D2FE">Total</th>`;
  r1 += '</tr>';
  // Row 2: sub-headers
  let r2 = '<tr>';
  P.years.forEach((y,i)=>{
    const bl = i>0?'border-left:2px solid #E2E8F0':'';
    Q_COLS.forEach((q,qi)=>{
      const bl2 = (qi===0&&i>0)?bl:'';
      r2 += `<th style="${bl2}">${q}</th>`;
    });
    r2 += `<th style="">Arts</th><th>Docs</th>`;
  });
  r2 += `<th style="background:#EFF6FF;border-left:2px solid #C7D2FE">Arts</th>`;
  r2 += `<th style="background:#EFF6FF">Docs</th>`;
  r2 += '</tr>';
  head.innerHTML = r1 + r2;
}

function renderPivot(){
  const filter = document.getElementById('pivotSchool')?.value || '';
  const rows   = P.rows.filter(r=> !filter || r.school === filter);
  const body   = document.getElementById('pivotBody');
  if(!body) return;

  let html = '';
  let lastSchool = null;
  let shownCount = 0;

  rows.forEach(r=>{
    // School group header
    if(r.school !== lastSchool){
      lastSchool = r.school;
      const totalCols = P.years.length * 7 + 2;
      html += `<tr class="school-hdr"><td class="c-fix-1" colspan="3" style="left:0;position:sticky;z-index:4">🏫 ${r.school}</td>`;
      for(let i=0; i<totalCols; i++) html += '<td></td>';
      html += '</tr>';
    }

    let grandArts = 0, grandDocs = 0;
    let cells = '';
    P.years.forEach((y,yi)=>{
      const yd    = r[y] || {};
      const bl    = yi > 0 ? 'border-left:2px solid #E2E8F0' : '';
      Q_COLS.forEach((q,qi)=>{
        const v   = yd[q] || 0;
        const css = v===0 ? 'zerv' : Q_CSS[q];
        const bl2 = (qi===0 && yi>0) ? `style="${bl}"` : '';
        cells += `<td class="${css}" ${bl2}>${v===0?'—':v}</td>`;
      });
      const arts = yd.total_arts || 0;
      const docs = yd.total_docs || 0;
      cells += `<td class="${arts===0?'zerv':'tot-art'}">${arts===0?'—':arts}</td>`;
      cells += `<td class="${docs===0?'zerv':'tot-doc'}">${docs===0?'—':docs}</td>`;
      grandArts += arts;
      grandDocs += docs;
    });

    html += `<tr>
      <td class="c-fix-1" title="${r.name}">${r.name}</td>
      <td class="c-fix-2">${r.school}</td>
      <td class="c-fix-3" style="font-family:monospace;font-size:11px">${r.scopus_id}</td>
      ${cells}
      <td class="${grandArts===0?'zerv':'tot-art'}" style="background:#F5F3FF;border-left:2px solid #C7D2FE;font-weight:700">${grandArts===0?'—':grandArts}</td>
      <td class="${grandDocs===0?'zerv':'tot-doc'}" style="background:#F5F3FF;font-weight:700">${grandDocs===0?'—':grandDocs}</td>
    </tr>`;
    shownCount++;
  });

  body.innerHTML = html;
  const stat = document.getElementById('pivot-stat');
  if(stat) stat.textContent = `${shownCount} autores mostrados`;
}

function setupPivotFilter(){
  const sel = document.getElementById('pivotSchool');
  if(!sel) return;
  const schools = [...new Set(P.rows.map(r=>r.school))].sort();
  schools.forEach(s=>{
    const opt = document.createElement('option');
    opt.value = opt.textContent = s;
    sel.appendChild(opt);
  });
}

// ── RENDER ALL (year-sensitive charts) ───────────────────────────
function renderAll(){
  updateKPIs();
  drawTypeDonut();
  drawQ1Q2Mini();
  drawQDonut();
  drawPctQ1();
  drawSchoolQ();
  drawSchoolType();
  drawAuthType();
  drawAuthQ1();
  drawPairs();
  drawAreaQ();
  drawAreaPct();
}

// ── YEAR FILTER SETUP ────────────────────────────────────────────
function setupYearFilter(){
  const sel = document.getElementById('yearFilter');
  D.meta.years.forEach(y=>{
    const opt = document.createElement('option');
    opt.value = opt.textContent = y;
    sel.appendChild(opt);
  });
  sel.addEventListener('change', e=>{ year = e.target.value; renderAll(); });
}

// ── NAV HIGHLIGHT ────────────────────────────────────────────────
function setupNav(){
  const links = document.querySelectorAll('.nav-link');
  const observer = new IntersectionObserver(entries=>{
    entries.forEach(en=>{
      if(en.isIntersecting){
        links.forEach(l=>l.classList.remove('active'));
        const active = document.querySelector(`.nav-link[href="#${en.target.id}"]`);
        if(active) active.classList.add('active');
      }
    });
  },{rootMargin:'-40% 0px -55% 0px'});
  document.querySelectorAll('section[id]').forEach(s=>observer.observe(s));
}

// ── INIT ─────────────────────────────────────────────────────────
document.addEventListener('DOMContentLoaded',()=>{
  setupYearFilter();
  setupNav();
  applyTheme(darkMode);  // restore saved theme
  buildPivotHeader();
  setupPivotFilter();
  renderPivot();
  drawTimeline();  // fixed (all years)
  drawQTrend();    // fixed (all years)
  renderAll();     // year-sensitive charts
});
</script>
</body>
</html>
"""

# ─── COMPUTE DYNAMIC HEIGHT HINTS ────────────────────────────────────────────
n_schools = len(by_year["ALL"]["schools"])
n_authors = TOP_AUTHORS
n_pairs   = min(TOP_PAIRS, len(by_year["ALL"]["pairs"]))
n_areas   = min(TOP_AREAS, len(by_year["ALL"]["areas"]))
sch_h   = max(200, n_schools * 34 + 80)
auth_h  = max(300, n_authors * 30 + 50)
pairs_h = max(300, n_pairs   * 30 + 50)
area_h  = max(300, n_areas   * 30 + 80)

# ─── SUBSTITUTE PLACEHOLDERS ─────────────────────────────────────────────────
OUT_DIR.mkdir(parents=True, exist_ok=True)
out_html = (HTML
    .replace("__DATA_JSON__",   DATA_JSON)
    .replace("__START_YEAR__",  str(START_YEAR))
    .replace("__ACTU__",        str(actu))
    .replace("__TOP_AUTHORS__", str(TOP_AUTHORS))
    .replace("__TOP_PAIRS__",   str(TOP_PAIRS))
    .replace("__TOP_AREAS__",   str(TOP_AREAS))
    .replace("__SCH_H__",       str(sch_h))
    .replace("__AUTH_H__",      str(auth_h))
    .replace("__PAIRS_H__",     str(pairs_h))
    .replace("__AREA_H__",      str(area_h))
)

out_path = OUT_DIR / "index.html"
out_path.write_text(out_html, encoding="utf-8")
size_kb = out_path.stat().st_size // 1024
print(f"\n✅  Dashboard written → {out_path}  ({size_kb} KB)")
print(f"   Sections: Producción · Calidad(Scimago) · Escuelas · Autores · Colaboración · Áreas · Tabla · Metodología")
print(f"   Charts: 13 gráficos Chart.js + tabla pivot interactiva")
print(f"   Height controls: 10 bar charts con slider ⇕")

# ─── EXCEL PIVOT ──────────────────────────────────────────────────────────────
print("\nGenerating Excel pivot...")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pivot Autores"

# ── Style helpers ──────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))
def _font(bold=False, color="1E293B", size=11):
    return Font(bold=bold, color=color.lstrip("#"), size=size, name="Calibri")
def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _border_thin(sides="all"):
    thin = Side(style="thin", color="E2E8F0")
    thick = Side(style="medium", color="94A3B8")
    b = {"left": thin, "right": thin, "top": thin, "bottom": thin}
    if sides == "left_thick":  b["left"]  = thick
    if sides == "top_thick":   b["top"]   = thick
    return Border(**b)

# Color palettes
HDR_FILL  = _fill("#F0F4F9")
YR_FILLS  = [_fill("#EFF6FF"), _fill("#F0FDF4"), _fill("#FEF9C3"),
             _fill("#FFF1F2"), _fill("#F5F3FF")]
Q_COLORS  = {"Q1":"059669","Q2":"2563EB","Q3":"D97706","Q4":"DC2626","SC":"94A3B8"}
TOT_FILL  = _fill("#F5F3FF")
SCH_FILL  = _fill("#DBEAFE")
ROW_EVEN  = _fill("#FAFBFC")

FIXED_COLS = 3  # Docente, Escuela, Scopus ID
Q_SUB      = ["Q1","Q2","Q3","Q4","SC","Arts","Docs"]  # 7 per year
TOTAL_COLS = FIXED_COLS + len(years_list) * len(Q_SUB) + 2  # +2 for grand total

# ── Row 1: top headers ─────────────────────────────────────────────────────────
row1 = [None, None, None]  # fixed cols (will merge vertically)
for yi, y in enumerate(years_list):
    row1 += [str(y)] + [None]*6   # 7 cols per year, merged
row1 += ["Total", None]           # grand total, merged

ws.append(row1)

# ── Row 2: sub-headers ────────────────────────────────────────────────────────
row2 = ["Docente", "Escuela", "Scopus ID"]
for y in years_list:
    row2 += Q_SUB
row2 += ["Arts", "Docs"]

ws.append(row2)

# ── Merge year-group headers ───────────────────────────────────────────────────
# Fixed 3 cols: merge rows 1-2
for c in range(1, FIXED_COLS+1):
    ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

# Year groups: merge across 7 cols in row 1
for yi in range(len(years_list)):
    c_start = FIXED_COLS + yi * len(Q_SUB) + 1
    c_end   = c_start + len(Q_SUB) - 1
    ws.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)

# Grand total: merge across 2 cols in row 1
tot_start = FIXED_COLS + len(years_list) * len(Q_SUB) + 1
ws.merge_cells(start_row=1, start_column=tot_start, end_row=1, end_column=tot_start+1)

# ── Style header rows ─────────────────────────────────────────────────────────
for r_idx in [1, 2]:
    for c_idx in range(1, TOTAL_COLS + 1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.font      = _font(bold=True, color="1E293B", size=11)
        cell.alignment = _align()
        cell.border    = _border_thin()
        # Year-group background
        col_in_data = c_idx - FIXED_COLS
        if col_in_data > 0:
            yi = (col_in_data - 1) // len(Q_SUB)
            if yi < len(years_list):
                cell.fill = YR_FILLS[yi % len(YR_FILLS)]
            else:
                cell.fill = TOT_FILL
        else:
            cell.fill = HDR_FILL
        # Sub-header Q column colors (row 2)
        if r_idx == 2 and col_in_data > 0:
            sub_idx = (col_in_data - 1) % len(Q_SUB)
            if sub_idx < 5:
                q = Q_SUB[sub_idx]
                cell.font = _font(bold=True, color=Q_COLORS[q], size=11)

# ── Data rows ─────────────────────────────────────────────────────────────────
current_school = None
data_row_start = 3
for r in authors_pivot:
    # School separator row
    if r["school"] != current_school:
        current_school = r["school"]
        sep_row = [f"  🏫  {current_school}"] + [""] * (TOTAL_COLS - 1)
        ws.append(sep_row)
        sr = ws.max_row
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=TOTAL_COLS)
        c = ws.cell(row=sr, column=1)
        c.fill      = SCH_FILL
        c.font      = _font(bold=True, color="1D4ED8", size=11)
        c.alignment = _align(h="left")
        c.border    = _border_thin("top_thick")

    # Author data row
    grand_arts = grand_docs = 0
    row_data = [r["name"], r["school"], r["scopus_id"]]
    for y in years_list:
        yd = r.get(str(y), {})
        for q in Q_SUB:
            row_data.append(yd.get(q, 0) if q in ["Q1","Q2","Q3","Q4","SC"]
                            else yd.get("total_arts",0) if q=="Arts"
                            else yd.get("total_docs",0))
        grand_arts += yd.get("total_arts", 0)
        grand_docs += yd.get("total_docs", 0)
    row_data += [grand_arts, grand_docs]
    ws.append(row_data)

    dr = ws.max_row
    is_even = (dr % 2 == 0)
    for c_idx in range(1, TOTAL_COLS + 1):
        cell     = ws.cell(row=dr, column=c_idx)
        col_data = c_idx - FIXED_COLS
        cell.border    = _border_thin()
        cell.alignment = _align(h="left" if c_idx <= FIXED_COLS else "center")
        # Background
        if c_idx <= FIXED_COLS:
            cell.fill = ROW_EVEN if is_even else PatternFill()
            cell.font = _font(bold=(c_idx==1), color="0F172A")
        elif col_data > len(years_list)*len(Q_SUB):
            cell.fill = TOT_FILL
            cell.font = _font(bold=True, color="4F46E5" if (col_data % 2 == 1) else "0891B2")
        else:
            yi      = (col_data - 1) // len(Q_SUB)
            sub_idx = (col_data - 1) % len(Q_SUB)
            cell.fill = YR_FILLS[yi % len(YR_FILLS)] if is_even else _fill("FFFFFF")
            v = cell.value or 0
            if sub_idx < 5 and v > 0:
                q = Q_SUB[sub_idx]
                cell.font = _font(bold=(sub_idx<4), color=Q_COLORS[q])
            elif v == 0:
                cell.value = "—"
                cell.font  = _font(color="CBD5E1")
            elif sub_idx == 5:
                cell.font = _font(bold=True, color="4F46E5")
            else:
                cell.font = _font(bold=False, color="0891B2")

# ── Column widths ──────────────────────────────────────────────────────────────
ws.column_dimensions[get_column_letter(1)].width = 28  # Docente
ws.column_dimensions[get_column_letter(2)].width = 22  # Escuela
ws.column_dimensions[get_column_letter(3)].width = 14  # Scopus ID
for c_idx in range(FIXED_COLS+1, TOTAL_COLS+1):
    ws.column_dimensions[get_column_letter(c_idx)].width = 7

# ── Freeze panes ──────────────────────────────────────────────────────────────
ws.freeze_panes = get_column_letter(FIXED_COLS+1) + "3"

# ── Row heights ───────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 18

# ── Save ──────────────────────────────────────────────────────────────────────
pivot_path = OUT_DIR / "autores_pivot_cuartiles.xlsx"
wb.save(pivot_path)
pivot_kb = pivot_path.stat().st_size // 1024
print(f"✅  Pivot Excel → {pivot_path}  ({pivot_kb} KB)")
print(f"   {len(authors_pivot)} autores · {len(years_list)} años · {TOTAL_COLS} columnas")
